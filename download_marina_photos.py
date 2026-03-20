"""
═══════════════════════════════════════════════════════════════════════════════
  YACHT HARBOURS — Marina Photo Downloader
  Pulls photos for every marina in your database from multiple sources
═══════════════════════════════════════════════════════════════════════════════

  SETUP:
    pip install requests openpyxl pillow

  USAGE:
    python3 download_marina_photos.py

  WHAT IT DOES:
    1. Reads your Marina Database Excel file
    2. For each marina with lat/lon, tries to fetch a photo from:
       a) Marinas.com API  (best quality, free, no key needed)
       b) Wikimedia Commons (CC-licensed, great for famous marinas)
       c) Mapbox Satellite  (aerial view — needs free Mapbox token, fallback)
    3. Downloads photos to  ./marina_photos/{marina_id}/
    4. Updates the Excel with Photo_URL and Photo_Status columns
    5. Generates a photo_report.html you can open to review all images

  TOKENS NEEDED:
    MAPBOX_TOKEN  — free at account.mapbox.com (50k images/month free)
    FLICKR_KEY    — free at flickr.com/services/api/ (optional)

═══════════════════════════════════════════════════════════════════════════════
"""

import os
import re
import json
import time
import hashlib
import requests
from pathlib import Path
from urllib.parse import urlencode, quote
from openpyxl import load_workbook

# ── CONFIG — set your tokens here ──────────────────────────────────────────
MAPBOX_TOKEN   = "YOUR_MAPBOX_TOKEN_HERE"   # account.mapbox.com — free
FLICKR_API_KEY = ""                          # optional
INPUT_XLSX     = "YachtHarbours_Marina_Database.xlsx"
PHOTO_DIR      = Path("./marina_photos")
REPORT_HTML    = Path("./photo_report.html")
DELAY_SECONDS  = 0.5   # be polite to APIs
MAX_MARINAS    = None   # set to 100 to test, None for all
# ───────────────────────────────────────────────────────────────────────────

HEADERS = {"User-Agent": "YachtHarbours/1.0 (www.yachtharbours.com)"}
session = requests.Session()
session.headers.update(HEADERS)

PHOTO_DIR.mkdir(exist_ok=True)

stats = {"total": 0, "marinas_com": 0, "wikimedia": 0, "mapbox": 0, "failed": 0}


def slugify(text):
    return re.sub(r'[^a-z0-9_-]', '_', str(text).lower())[:60]


# ── SOURCE 1: Marinas.com API ───────────────────────────────────────────────
def fetch_marinas_com(marina_id, name, lat, lon):
    """Search Marinas.com API by location, match by name, return best image URL."""
    try:
        url = "https://api.marinas.com/v1/points/search"
        params = {"location[lat]": lat, "location[lon]": lon}
        r = session.get(url, params=params, timeout=10)
        if r.status_code != 200:
            return None
        data = r.json().get("data", [])
        name_lower = name.lower()
        for item in data:
            item_name = item.get("name", "").lower()
            # Fuzzy match: check if significant words overlap
            words = [w for w in name_lower.split() if len(w) > 3]
            if any(w in item_name for w in words):
                imgs = item.get("images", {}).get("data", [])
                if imgs:
                    return imgs[0].get("medium_url") or imgs[0].get("thumbnail_url")
        # fallback — just take first result if very close
        if data:
            imgs = data[0].get("images", {}).get("data", [])
            if imgs:
                return imgs[0].get("medium_url")
    except Exception as e:
        pass
    return None


# ── SOURCE 2: Wikimedia Commons ────────────────────────────────────────────
def fetch_wikimedia(name, city, country):
    """Search Wikimedia Commons for a CC-licensed marina photo."""
    queries = [
        f"{name} marina",
        f"{city} marina harbour",
        f"{name} harbour {country}",
    ]
    for q in queries:
        try:
            url = "https://commons.wikimedia.org/w/api.php"
            params = {
                "action": "query", "format": "json",
                "generator": "search", "gsrsearch": q,
                "gsrnamespace": "6", "gsrlimit": "5",
                "prop": "imageinfo",
                "iiprop": "url|extmetadata",
                "iiurlwidth": 600,
            }
            r = session.get(url, params=params, timeout=10)
            if r.status_code != 200:
                continue
            pages = r.json().get("query", {}).get("pages", {})
            for _, page in pages.items():
                ii = page.get("imageinfo", [{}])[0]
                meta = ii.get("extmetadata", {})
                license_short = meta.get("LicenseShortName", {}).get("value", "")
                # Only use freely licensed images
                if any(x in license_short.upper() for x in ["CC", "PUBLIC", "CC0"]):
                    thumb = ii.get("thumburl")
                    if thumb:
                        return thumb
        except Exception:
            pass
        time.sleep(DELAY_SECONDS)
    return None


# ── SOURCE 3: Mapbox Satellite (fallback — always works with token) ─────────
def get_mapbox_url(lat, lon, zoom=15, width=600, height=400):
    """Generate Mapbox Satellite Static Image URL (no download — used as src in HTML)."""
    if not MAPBOX_TOKEN or MAPBOX_TOKEN == "YOUR_MAPBOX_TOKEN_HERE":
        return None
    return (f"https://api.mapbox.com/styles/v1/mapbox/satellite-v9/static/"
            f"{lon},{lat},{zoom},0/{width}x{height}@2x"
            f"?access_token={MAPBOX_TOKEN}")


# ── DOWNLOAD helper ─────────────────────────────────────────────────────────
def download_image(url, dest_path):
    """Download image to dest_path. Returns True on success."""
    try:
        r = session.get(url, timeout=20, stream=True)
        if r.status_code == 200 and "image" in r.headers.get("content-type", ""):
            with open(dest_path, "wb") as f:
                for chunk in r.iter_content(8192):
                    f.write(chunk)
            return True
    except Exception:
        pass
    return False


# ── MAIN ─────────────────────────────────────────────────────────────────────
def process_marina(marina_id, name, city, country, lat, lon):
    """Try each source in order until we get a photo. Returns (local_path, source, url)."""
    folder = PHOTO_DIR / slugify(marina_id)
    folder.mkdir(exist_ok=True)

    # Try Marinas.com first
    img_url = fetch_marinas_com(marina_id, name, lat, lon)
    if img_url:
        dest = folder / "photo_1.jpg"
        if download_image(img_url, dest):
            return str(dest), "Marinas.com", img_url
    time.sleep(DELAY_SECONDS)

    # Try Wikimedia Commons
    img_url = fetch_wikimedia(name, city, country)
    if img_url:
        dest = folder / "photo_1.jpg"
        if download_image(img_url, dest):
            return str(dest), "Wikimedia Commons", img_url
    time.sleep(DELAY_SECONDS)

    # Mapbox satellite (URL only — not downloaded, used as <img src>)
    mapbox_url = get_mapbox_url(lat, lon)
    if mapbox_url:
        return f"MAPBOX_URL:{mapbox_url}", "Mapbox Satellite", mapbox_url

    return None, "Not Found", None


def run():
    if not Path(INPUT_XLSX).exists():
        print(f"❌ Input file not found: {INPUT_XLSX}")
        print("   Please run from same folder as your Marina Database Excel file.")
        return

    print("═" * 65)
    print("  YACHT HARBOURS — Marina Photo Downloader")
    print("═" * 65)

    wb = load_workbook(INPUT_XLSX)
    ws = wb.active

    # Find column indices
    header_row = {ws.cell(2, c).value: c for c in range(1, ws.max_column + 1)}
    col_id   = header_row.get("ID", 1)
    col_name = header_row.get("Marina Name", 2)
    col_city = header_row.get("City / Town", 7)
    col_ctry = header_row.get("Country", 5)
    col_lat  = header_row.get("Latitude", 8)
    col_lon  = header_row.get("Longitude", 9)

    # Add photo columns if not present
    last_col = ws.max_column + 1
    if "Photo URL" not in header_row:
        ws.cell(2, last_col, "Photo URL")
        ws.cell(2, last_col + 1, "Photo Source")
        ws.cell(2, last_col + 2, "Photo Status")
        col_photo_url    = last_col
        col_photo_source = last_col + 1
        col_photo_status = last_col + 2
    else:
        col_photo_url    = header_row["Photo URL"]
        col_photo_source = header_row.get("Photo Source", last_col)
        col_photo_status = header_row.get("Photo Status", last_col + 1)

    report_rows = []

    for row in range(3, ws.max_row + 1):
        if MAX_MARINAS and stats["total"] >= MAX_MARINAS:
            break

        marina_id = ws.cell(row, col_id).value
        name      = ws.cell(row, col_name).value
        lat       = ws.cell(row, col_lat).value
        lon       = ws.cell(row, col_lon).value
        city      = ws.cell(row, col_city).value or ""
        country   = ws.cell(row, col_ctry).value or ""

        if not all([marina_id, name, lat, lon]):
            continue

        stats["total"] += 1
        print(f"[{stats['total']:>4}] {name[:40]:<40}", end=" ", flush=True)

        local_path, source, img_url = process_marina(
            marina_id, name, city, country, float(lat), float(lon))

        if local_path:
            ws.cell(row, col_photo_url).value = img_url
            ws.cell(row, col_photo_source).value = source
            ws.cell(row, col_photo_status).value = "✅ Found"
            if source == "Marinas.com":  stats["marinas_com"] += 1
            elif source == "Wikimedia":  stats["wikimedia"]   += 1
            elif source == "Mapbox Satellite": stats["mapbox"] += 1
            print(f"✅ {source}")
        else:
            ws.cell(row, col_photo_status).value = "❌ Not Found"
            stats["failed"] += 1
            print("❌ Not found")

        report_rows.append((marina_id, name, city, country, lat, lon,
                            img_url, source, local_path))

    wb.save(INPUT_XLSX.replace(".xlsx", "_with_photos.xlsx"))
    print(f"\n✅ Updated database saved")

    # Generate HTML report
    html_rows = ""
    for mid, nm, cty, ctry, lat, lon, url, src, path in report_rows:
        if url:
            img_src = url if url.startswith("http") else url.replace("MAPBOX_URL:", "")
            status_badge = f'<span style="color:green">✅ {src}</span>'
        else:
            img_src = f"https://api.mapbox.com/styles/v1/mapbox/satellite-v9/static/{lon},{lat},14,0/320x200@2x?access_token={MAPBOX_TOKEN}"
            status_badge = '<span style="color:red">❌ No photo</span>'

        html_rows += f"""
        <div style="display:inline-block;width:300px;margin:8px;padding:10px;
                    border:1px solid #ddd;border-radius:8px;vertical-align:top;
                    font-family:Arial,sans-serif;font-size:12px">
          <img src="{img_src}" style="width:280px;height:160px;object-fit:cover;
               border-radius:4px;display:block;margin-bottom:6px"
               onerror="this.src='https://via.placeholder.com/280x160?text=No+Image'"
               loading="lazy">
          <strong style="color:#0A1628">{nm}</strong><br>
          <span style="color:#64748b">{cty}, {ctry}</span><br>
          {status_badge}
        </div>"""

    REPORT_HTML.write_text(f"""<!DOCTYPE html>
<html><head><title>Marina Photo Report — Yacht Harbours</title>
<style>body{{background:#f0f4f8;padding:20px;font-family:Arial}}
h1{{color:#0A1628}} .stats{{background:#0A1628;color:#C9A84C;
padding:16px;border-radius:8px;margin-bottom:20px}}</style></head>
<body>
<h1>🌊 Yacht Harbours — Marina Photo Report</h1>
<div class="stats">
  Total processed: {stats['total']} &nbsp;|&nbsp;
  Marinas.com: {stats['marinas_com']} &nbsp;|&nbsp;
  Wikimedia: {stats['wikimedia']} &nbsp;|&nbsp;
  Mapbox: {stats['mapbox']} &nbsp;|&nbsp;
  Not found: {stats['failed']}
</div>
{html_rows}
</body></html>""")

    print(f"\n📊 RESULTS:")
    print(f"   Total marinas processed : {stats['total']}")
    print(f"   Photos from Marinas.com : {stats['marinas_com']}")
    print(f"   Photos from Wikimedia   : {stats['wikimedia']}")
    print(f"   Mapbox satellite URLs   : {stats['mapbox']}")
    print(f"   Not found               : {stats['failed']}")
    print(f"\n📄 Open {REPORT_HTML} in your browser to review all images")
    print(f"📊 Updated Excel saved as YachtHarbours_Marina_Database_with_photos.xlsx")


if __name__ == "__main__":
    run()
